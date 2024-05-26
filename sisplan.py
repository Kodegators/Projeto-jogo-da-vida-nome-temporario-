from openpyxl import *
import random
from points import *
from Buttongeneric import *

class CurrentQuestion:
    def __init__(self):
        self.he = 0
        self.ha = 0
        self.re = 0
        self.mo = 0
        self.score = 0
        self.index = 0
        self.questions = []
        self.init_questions = []
        self.mid_questions = []
        self.end_questions = []
        
    #Pega 3 perguntas da tabela infância
    def questgeninit(self):
        self.init_questions = self.get_random_questions(self.sheetinit)
        self.questions.extend(self.init_questions)
    
    #Pega 3 perguntas da tabela adulto
    def questgenmid(self):
        self.mid_questions = self.get_random_questions(self.sheetmid)
        self.questions.extend(self.mid_questions)
        
    def questgenend(self):
        self.end_questions = self.final_quest(self.sheetend)
        self.questions.extend(self.end_questions)

    #Pegando perguntas aleatórias
    def get_random_questions(self,sheet):
        used_lines = set()
        questions = []
        while len(questions) <= 3:
            line = str(random.randint(2,9))
            if line not in used_lines:
                used_lines.add(line)
                cellquest = sheet['A'+line].value
                cellansw1 = sheet['B'+line].value
                cellansw2 = sheet['D'+line].value
                questions.append((line, cellquest, cellansw1, cellansw2))
            return questions
            
    def final_quest(self,sheet):
        questions = []
        for line in range(4):
            linend = line + 2
            cellquest = sheet['A'+str(linend)].value
            cellansw1 = sheet['B'+str(linend)].value
            cellansw2 = sheet['D'+str(linend)].value
            questions.append((str(linend), cellquest, cellansw1, cellansw2))    
        return questions

    #Pega a próxima pergunta da lista
    def get_next_question(self):
        if  self.index < len(self.questions):
            question = self.questions[self.index]
            self.index += 1
            return question 
        '''
        elif self.indexend < len(self.questions):
            questionend = self.questions[self.indexend]
            self.indexend += 1
            return questionend
        '''    

    #Carrega tabela infantil
    def load_init_table(self, filepath):
        self.tableinit = load_workbook(filepath)
        self.sheetinit = self.tableinit.active
        self.questgeninit()
        
    #Carrega tabela adulto
    def load_mid_table(self, filepath):
        self.tablemid = load_workbook(filepath)
        self.sheetmid = self.tablemid.active 
        self.questgenmid()

    def load_end_table(self, filepath):
        self.tableend = load_workbook(filepath)
        self.sheetend = self.tableend.active
        self.questgenend()
    
             